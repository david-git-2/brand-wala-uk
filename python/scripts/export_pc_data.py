import os
import re
import json
import time
from collections import defaultdict
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl

from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow


ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
DEFAULT_XLSX = os.path.join(ROOT_DIR, "python", "data", "pc_data.xlsx")
DEFAULT_OUT_JSON = os.path.join(ROOT_DIR, "web", "public", "data", "pc_data.json")
DEFAULT_OUT_IMAGES = os.path.join(ROOT_DIR, "python", "images", "out_images")

CREDS_DIR = os.path.join(ROOT_DIR, "python", "credentials")
OAUTH_CLIENT_JSON = os.path.join(CREDS_DIR, "oauth_client.json")
TOKEN_JSON = os.path.join(CREDS_DIR, "token.json")

SCOPES = ["https://www.googleapis.com/auth/drive.file"]  # only files this app creates


def log(msg: str):
    print(msg, flush=True)


def safe_filename(s: str) -> str:
    s = str(s).strip()
    s = re.sub(r"[^\w\-\.]+", "_", s)
    return s[:120] if len(s) > 120 else s


def prompt_int(label: str, default: int) -> int:
    """Prompt user for an integer. If empty input, returns default."""
    while True:
        raw = input(f"{label} [{default}]: ").strip()
        if raw == "":
            return default
        try:
            val = int(raw)
            if val <= 0:
                log("❌ Please enter a positive integer.")
                continue
            return val
        except ValueError:
            log("❌ Invalid number. Try again.")


def normalize_header(h: str) -> str:
    # normalize for matching required columns
    return re.sub(r"[\s\-]+", "_", str(h).strip().lower())


def get_oauth_credentials() -> Credentials:
    """
    Loads token.json if present, otherwise performs OAuth login.
    Refreshes expired tokens and saves back to token.json.
    Returns a valid Credentials object.
    """
    os.makedirs(CREDS_DIR, exist_ok=True)

    creds = None
    if os.path.exists(TOKEN_JSON):
        log("🔐 Found token.json, using existing login...")
        creds = Credentials.from_authorized_user_file(TOKEN_JSON, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            log("🔄 Token expired, refreshing...")
            creds.refresh(Request())
        else:
            if not os.path.exists(OAUTH_CLIENT_JSON):
                raise FileNotFoundError(
                    f"Missing OAuth client file: {OAUTH_CLIENT_JSON}\n"
                    f"Download 'Desktop app' OAuth client JSON and save it there."
                )
            log("🌐 Opening browser for Google login/consent...")
            flow = InstalledAppFlow.from_client_secrets_file(OAUTH_CLIENT_JSON, SCOPES)
            creds = flow.run_local_server(port=0)

        with open(TOKEN_JSON, "w", encoding="utf-8") as token:
            token.write(creds.to_json())
        log(f"✅ Saved token: {TOKEN_JSON}")

    return creds


def build_drive_service(creds: Credentials):
    """Build a Drive API client."""
    return build("drive", "v3", credentials=creds)


def upload_to_drive(service, local_path: str, folder_id: str) -> str:
    """
    Uploads a file to Drive into folder_id.
    IMPORTANT: Does NOT set per-file public permissions.
    You said your target folder is already public, so we skip permissions for speed.
    """
    filename = os.path.basename(local_path)
    metadata = {"name": filename}
    if folder_id:
        metadata["parents"] = [folder_id]

    media = MediaFileUpload(local_path, resumable=True)
    created = service.files().create(body=metadata, media_body=media, fields="id").execute()
    file_id = created["id"]

    # Direct link works well for <img src=""> if the file is accessible via folder sharing.
    direct_link = f"https://drive.google.com/uc?id={file_id}"
    return direct_link


def format_duration(seconds: float) -> str:
    seconds = max(0, int(seconds))
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    if h > 0:
        return f"{h}h {m}m {s}s"
    if m > 0:
        return f"{m}m {s}s"
    return f"{s}s"


def main():
    t0 = time.perf_counter()

    # ---- CONFIG ----
    EXCEL_PATH = DEFAULT_XLSX
    OUT_JSON_PATH = DEFAULT_OUT_JSON
    OUT_IMAGES_DIR = DEFAULT_OUT_IMAGES

    # Your Drive folder is already public (Anyone with link).
    DRIVE_FOLDER_ID = "1s-DCoV7rkhLllBhTVOm2SZ7IlA0JmiEB"

    DEFAULT_HEADER_ROW = 4
    DEFAULT_IMAGE_COLUMN_INDEX = 14  # 1-based column index (14 = N)

    SHEET_NAME = None

    # Parallel upload settings
    MAX_WORKERS = 6  # try 5-10; too high may hit rate limits
    # ----------------

    log(
        "\n"
        "🚀 PC export starting...\n"
        "\n"
        "📌 Excel format requirements:\n"
        "  Required columns in the HEADER ROW:\n"
        "    - product_code\n"
        "    - barcode\n"
        "    - case_size\n"
        "    - name\n"
        "    - price\n"
        "    - image   (images are embedded in the sheet; you will enter the image column index)\n"
        "  Optional columns:\n"
        "    - country_of_origin\n"
        "    - brand\n"
        "\n"
        "🆔 Product ID:\n"
        "  - product_id = product_code + '_' + barcode\n"
        "\n"
        "⚡ Speed mode enabled:\n"
        "  - Your Drive folder is already public → we will NOT set per-file permissions.\n"
        f"  - Parallel uploads enabled (workers={MAX_WORKERS}).\n"
        "\n"
        "👉 You will be asked for:\n"
        "  - Header row number (where the column names are)\n"
        "  - Image column index (1=A, 2=B, ...)\n"
    )

    HEADER_ROW = prompt_int("Enter header row number", DEFAULT_HEADER_ROW)
    IMAGE_COLUMN_INDEX = prompt_int("Enter image column index (1=A, 2=B, ...)", DEFAULT_IMAGE_COLUMN_INDEX)

    log(f"\n📄 Excel: {EXCEL_PATH}")
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel not found: {EXCEL_PATH}")

    os.makedirs(OUT_IMAGES_DIR, exist_ok=True)
    os.makedirs(os.path.dirname(OUT_JSON_PATH), exist_ok=True)
    log(f"📁 Output images folder: {OUT_IMAGES_DIR}")
    log(f"🧾 Output JSON: {OUT_JSON_PATH}\n")

    log("📥 Loading workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sh = wb[SHEET_NAME] if SHEET_NAME else wb[wb.sheetnames[0]]
    log(f"✅ Using sheet: {sh.title}")

    max_col = sh.max_column
    max_row = sh.max_row
    log(f"📐 Sheet size: rows={max_row}, cols={max_col}")

    # Read headers
    log(f"🏷️ Reading headers from row {HEADER_ROW}...")
    headers = []
    for c in range(1, max_col + 1):
        v = sh.cell(HEADER_ROW, c).value
        headers.append(str(v).strip() if v is not None else f"col_{c}")

    # Build header lookup (normalized -> column index)
    header_to_col = {}
    for idx, h in enumerate(headers, start=1):
        nh = normalize_header(h)
        header_to_col.setdefault(nh, idx)

    # Required/optional columns
    required = ["product_code", "barcode", "case_size", "name", "price"]
    optional = ["country_of_origin", "brand"]

    missing_required = [c for c in required if c not in header_to_col]
    if missing_required:
        raise RuntimeError(
            "❌ Missing required column(s) in header row "
            f"{HEADER_ROW}: {', '.join(missing_required)}\n"
            "Make sure your Excel header row contains these columns (case-insensitive):\n"
            "product_code, barcode, case_size, name, price\n"
        )

    missing_optional = [c for c in optional if c not in header_to_col]
    if missing_optional:
        log(f"ℹ️ Optional column(s) missing (OK): {', '.join(missing_optional)}")

    product_code_col = header_to_col["product_code"]
    barcode_col = header_to_col["barcode"]
    log(f"✅ Found product_code column at index: {product_code_col}")
    log(f"✅ Found barcode column at index: {barcode_col}")
    log(f"🖼️ Using image column index: {IMAGE_COLUMN_INDEX}\n")

    # Use actual header names (original casing/spaces) for row dict lookups
    product_code_header_name = headers[product_code_col - 1]
    barcode_header_name = headers[barcode_col - 1]

    # Read rows
    log("📦 Reading product rows...")
    start_data_row = HEADER_ROW + 1
    products_by_row = {}

    for r in range(start_data_row, max_row + 1):
        row_vals = [sh.cell(r, c).value for c in range(1, max_col + 1)]
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue

        obj = {}
        for c, h in enumerate(headers, start=1):
            obj[h] = row_vals[c - 1]
        obj["_rowNumber"] = r
        products_by_row[r] = obj

    log(f"✅ Products loaded: {len(products_by_row)}")

    # Extract images
    log(f"🖼️ Finding embedded images anchored to column {IMAGE_COLUMN_INDEX}...")
    images = getattr(sh, "_images", [])
    log(f"🖼️ Total images detected in sheet: {len(images)}")

    img_by_row = {}
    for img in images:
        anchor = img.anchor._from  # 0-based
        row = anchor.row + 1
        col = anchor.col + 1

        if col != IMAGE_COLUMN_INDEX:
            continue
        if row not in products_by_row:
            continue

        img_bytes = img._data()
        ext = (getattr(img, "format", None) or "jpg").lower()
        if ext == "jpeg":
            ext = "jpg"

        pc_val = products_by_row[row].get(product_code_header_name, "")
        bc_val = products_by_row[row].get(barcode_header_name, "")

        pc_str = safe_filename(pc_val if pc_val is not None else "NO_CODE")
        bc_str = safe_filename(bc_val if bc_val is not None else f"row_{row}")

        # Use product_code + barcode for uniqueness
        img_key = f"{pc_str}_{bc_str}".strip("_")
        img_by_row[row] = (img_key, ext, img_bytes)

    log(f"✅ Images matched to product rows: {len(img_by_row)}")

    # Save images locally
    log("💾 Saving images locally...")
    used = defaultdict(int)
    local_path_by_row = {}

    for i, (row, (img_key, ext, img_bytes)) in enumerate(img_by_row.items(), start=1):
        used[img_key] += 1
        suffix = f"_{used[img_key]}" if used[img_key] > 1 else ""
        filename = f"{img_key}{suffix}.{ext}"
        local_path = os.path.join(OUT_IMAGES_DIR, filename)

        with open(local_path, "wb") as f:
            f.write(img_bytes)

        local_path_by_row[row] = local_path
        if i % 25 == 0 or i == len(img_by_row):
            log(f"   ...saved {i}/{len(img_by_row)}")

    log(f"✅ Local images saved: {len(local_path_by_row)}\n")

    # OAuth once (IMPORTANT): do NOT do OAuth inside threads.
    log("☁️ Preparing Google Drive credentials...")
    base_creds = get_oauth_credentials()
    log("✅ Credentials ready")

    # Helper: create independent creds per worker to avoid shared-state issues
    base_creds_info = json.loads(base_creds.to_json())

    def upload_one(row: int, path: str):
        # Fresh creds object per thread (avoid races on refresh state)
        creds = Credentials.from_authorized_user_info(base_creds_info, SCOPES)
        service = build_drive_service(creds)
        url = upload_to_drive(service, path, DRIVE_FOLDER_ID)
        return row, url

    # Parallel upload
    items = list(local_path_by_row.items())
    total = len(items)
    drive_url_by_row = {}

    log(f"⬆️ Uploading {total} image(s) to Drive (parallel workers={MAX_WORKERS})...")
    up_start = time.perf_counter()

    done = 0
    failed = 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = [ex.submit(upload_one, row, path) for row, path in items]

        for fut in as_completed(futures):
            try:
                row, url = fut.result()
                drive_url_by_row[row] = url
            except Exception as e:
                failed += 1
                log(f"❌ Upload failed: {e}")
            finally:
                done += 1

                if done % 25 == 0 or done == total:
                    elapsed = time.perf_counter() - up_start
                    rate = (done / elapsed) if elapsed > 0 else 0.0
                    remaining = total - done
                    eta = (remaining / rate) if rate > 0 else 0
                    log(
                        f"   ...uploaded {done}/{total} "
                        f"(fail={failed}) | avg {rate:.2f} files/sec | ETA {format_duration(eta)}"
                    )

    uploaded_count = len(drive_url_by_row)
    log(f"✅ Upload step done. URLs created: {uploaded_count} (failed={failed})\n")

    # Build JSON (field names come from Excel header row)
    log("🧾 Building JSON payload...")
    products = []

    for row, obj in products_by_row.items():
        out = dict(obj)

        pc_val = obj.get(product_code_header_name, "")
        bc_val = obj.get(barcode_header_name, "")

        pc_str = safe_filename(pc_val if pc_val is not None else "")
        bc_str = safe_filename(bc_val if bc_val is not None else "")

        # ✅ stable product id
        out["product_id"] = f"{pc_str}_{bc_str}".strip("_")

        # ✅ imageUrl (drive direct link)
        out["imageUrl"] = drive_url_by_row.get(row)

        products.append(out)

    payload = {
        "meta": {
            "generatedAt": datetime.utcnow().isoformat() + "Z",
            "sourceFile": os.path.basename(EXCEL_PATH),
            "sheet": sh.title,
            "count": len(products),
            "imagesExtracted": len(local_path_by_row),
            "imagesUploaded": uploaded_count,
            "headerRow": HEADER_ROW,
            "imageColumnIndex": IMAGE_COLUMN_INDEX,
            "parallelWorkers": MAX_WORKERS,
            "driveFolderId": DRIVE_FOLDER_ID,
            "note": "Per-file permissions not set (folder is already public).",
            "productIdRule": "product_id = product_code + '_' + barcode",
        },
        "products": products,
    }

    with open(OUT_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    t_total = time.perf_counter() - t0

    log("\n✅ Done")
    log(f"- Products: {len(products)}")
    log(f"- Images extracted: {len(local_path_by_row)} -> {OUT_IMAGES_DIR}")
    log(f"- Images uploaded (with URL): {uploaded_count}")
    log(f"- Upload failures: {failed}")
    log(f"- JSON written: {OUT_JSON_PATH}")
    log(f"- Token saved: {TOKEN_JSON}")
    log(f"⏱️ Total time: {format_duration(t_total)}\n")


if __name__ == "__main__":
    main()